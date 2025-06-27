"""
Creates an Excel trading journal template with formulas and formatting
Run this to generate your Excel journal
"""

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    import datetime

    def create_excel_journal():
        wb = openpyxl.Workbook()

        # Sheet 1: Trade Log
        ws1 = wb.active
        ws1.title = "Trade Log"

        # Headers
        headers = [
            "Date", "Time", "Symbol", "Long/Short", "Strategy",
            "Entry Price", "Stop Loss", "Target", "Exit Price",
            "Shares", "Gross P&L", "Commissions", "Net P&L",
            "R-Multiple", "Win/Loss", "Account Balance", "Notes"
        ]

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Add formulas for row 2 as example
        ws1["K2"] = "=(I2-F2)*J2"  # Gross P&L
        ws1["M2"] = "=K2-L2"  # Net P&L
        ws1["N2"] = "=IF(F2>G2,(I2-F2)/(F2-G2),(I2-F2)/(G2-F2))"  # R-Multiple
        ws1["O2"] = '=IF(M2>0,"Win","Loss")'  # Win/Loss
        ws1["P2"] = "=P1+M2"  # Running balance

        # Sheet 2: Statistics
        ws2 = wb.create_sheet("Statistics")

        stats_layout = [
            ["TRADING STATISTICS", ""],
            ["", ""],
            ["Total Trades:", "=COUNTA('Trade Log'!A:A)-1"],
            ["Winning Trades:", "=COUNTIF('Trade Log'!O:O,\"Win\")"],
            ["Losing Trades:", "=COUNTIF('Trade Log'!O:O,\"Loss\")"],
            ["Win Rate:", "=IF(B3>0,B4/B3,0)"],
            ["", ""],
            ["Total P&L:", "=SUM('Trade Log'!M:M)"],
            ["Average Win:", "=AVERAGEIF('Trade Log'!O:O,\"Win\",'Trade Log'!M:M)"],
            ["Average Loss:", "=AVERAGEIF('Trade Log'!O:O,\"Loss\",'Trade Log'!M:M)"],
            ["Profit Factor:", "=IF(B10<0,ABS(B9/B10),0)"],
            ["", ""],
            ["Largest Win:", "=MAX('Trade Log'!M:M)"],
            ["Largest Loss:", "=MIN('Trade Log'!M:M)"],
            ["Average R:", "=AVERAGE('Trade Log'!N:N)"],
            ["", ""],
            ["Best Day:", "=MAX(SUMIF('Trade Log'!A:A,'Trade Log'!A:A,'Trade Log'!M:M))"],
            ["Worst Day:", "=MIN(SUMIF('Trade Log'!A:A,'Trade Log'!A:A,'Trade Log'!M:M))"],
        ]

        for row_idx, row_data in enumerate(stats_layout, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws2.cell(row=row_idx, column=col_idx, value=value)

        # Format statistics
        title_cell = ws2["A1"]
        title_cell.font = Font(bold=True, size=14)

        # Sheet 3: Monthly Summary
        ws3 = wb.create_sheet("Monthly Summary")
        ws3["A1"] = "Month"
        ws3["B1"] = "Total Trades"
        ws3["C1"] = "Win Rate"
        ws3["D1"] = "Net P&L"
        ws3["E1"] = "Best Trade"
        ws3["F1"] = "Worst Trade"

        # Save the workbook
        wb.save("Trading_Journal_Template.xlsx")
        print("Excel journal template created successfully!")
        return True

except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    return False

if __name__ == "__main__":
    create_excel_journal()
else:
    print("This script is intended to be run directly to create the Excel journal template.")
    print("Import it in another script if you want to use the create_excel_journal function.")
# This code creates an Excel trading journal template with formulas and formatting.
# It includes a trade log, statistics, and monthly summary sheets.
